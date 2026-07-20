"""
Sample Data Generator - サンプルExcelファイルを生成するスクリプト
"""

if __name__ == '__main__':
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # ワークブック作成
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tasks'
        
        # スタイル定義
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行
        headers = ['業務テーマ', '成果物', 'タスク', '開始時期', '終了時期', 'ステータス']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # サンプルデータ
        data = [
            ('システム基盤構築', None, None, '2026-01-15', '2026-04-30', None),
            (None, 'サーバ構築', None, '2026-01-15', '2026-02-28', None),
            (None, None, 'OS インストール', '2026-01-15', '2026-01-31', '完了'),
            (None, None, 'ネットワーク設定', '2026-02-01', '2026-02-28', '完了'),
            (None, 'データベース構築', None, '2026-03-01', '2026-04-30', None),
            (None, None, 'スキーマ設計', '2026-03-01', '2026-03-31', '完了'),
            (None, None, 'DB インスタンス構築', '2026-04-01', '2026-04-30', '着手中'),
            ('データ移行', None, None, '2026-03-01', '2026-06-30', None),
            (None, 'レガシーシステム抽出', None, '2026-03-01', '2026-04-30', None),
            (None, None, 'データ抽出', '2026-03-01', '2026-04-15', '着手中'),
            (None, None, 'データ検証', '2026-04-16', '2026-04-30', '未着手'),
            (None, '新システム投入', None, '2026-05-01', '2026-06-30', None),
            (None, None, 'データ変換', '2026-05-01', '2026-06-15', '未着手'),
            (None, None, 'データロード', '2026-06-16', '2026-06-30', '未着手'),
            ('セキュリティ構築', None, None, '2026-02-01', '2026-05-31', None),
            (None, 'セキュリティ設計', None, '2026-02-01', '2026-03-15', None),
            (None, None, 'セキュリティポリシー策定', '2026-02-01', '2026-02-28', '完了'),
            (None, None, 'リスク分析', '2026-03-01', '2026-03-15', '着手中'),
            (None, 'セキュリティ実装', None, '2026-03-16', '2026-05-31', None),
            (None, None, 'ファイアウォール構築', '2026-03-16', '2026-04-30', '着手中'),
            (None, None, '認証システム導入', '2026-05-01', '2026-05-31', '未着手'),
        ]
        
        # データの書き込み
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                
                # ステータス列のスタイル
                if col_idx == 6 and value is not None:
                    if value == '完了':
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        cell.font = Font(color='006100')
                    elif value == '着手中':
                        cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                        cell.font = Font(color='9C6500')
                    elif value == '未着手':
                        cell.fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
                        cell.font = Font(color='9C5700')
                
                # 日付列のアライメント
                if col_idx in (4, 5):
                    cell.alignment = Alignment(horizontal='center')
        
        # 列幅の調整
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12
        
        # ファイルの保存
        wb.save('sample.xlsx')
        print('✓ サンプルExcelファイルを生成しました: sample.xlsx')
        
    except ImportError as e:
        print(f'エラー: 必要なパッケージがインストールされていません: {e}')
        print('実行してください: pip install openpyxl')
